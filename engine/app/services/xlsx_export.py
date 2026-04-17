"""Tabular exports to XLSX for STR / alert / dissemination / entity lists.

Keeps the per-domain column layout inline so each domain exports its
native columns. The service reuses the existing list services (same
filters as the list routes) so an export mirrors whatever the user
sees on their screen.
"""
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _write_sheet(ws: Any, title: str, columns: list[tuple[str, str]], rows: Iterable[dict[str, Any]]) -> None:
    ws.title = title[:31]  # Excel caps sheet names at 31 chars
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="left", vertical="center")

    headers = [header for header, _ in columns]
    ws.append(headers)
    for idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(headers[idx - 1]) + 4)

    for row in rows:
        ws.append([row.get(key, "") for _, key in columns])


def _finalize(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_str_reports_xlsx(reports: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    columns: list[tuple[str, str]] = [
        ("Report Ref", "report_ref"),
        ("Type", "report_type"),
        ("Status", "status"),
        ("Organization", "org_name"),
        ("Subject Name", "subject_name"),
        ("Subject Account", "subject_account"),
        ("Subject Bank", "subject_bank"),
        ("Category", "category"),
        ("Total Amount", "total_amount"),
        ("Currency", "currency"),
        ("Transactions", "transaction_count"),
        ("Primary Channel", "primary_channel"),
        ("Risk Score", "auto_risk_score"),
        ("Cross-Bank Hit", "cross_bank_hit"),
        ("Reported At", "reported_at"),
    ]
    _write_sheet(ws, "Reports", columns, reports)
    return _finalize(wb)


def build_alerts_xlsx(alerts: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    columns: list[tuple[str, str]] = [
        ("Alert Title", "title"),
        ("Alert Type", "alert_type"),
        ("Severity", "severity"),
        ("Risk Score", "risk_score"),
        ("Status", "status"),
        ("Organization", "org_name"),
        ("Assigned To", "assigned_to"),
        ("Source", "source_type"),
        ("Created At", "created_at"),
    ]
    _write_sheet(ws, "Alerts", columns, alerts)
    return _finalize(wb)


def build_disseminations_xlsx(records: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    columns: list[tuple[str, str]] = [
        ("Dissemination Ref", "dissemination_ref"),
        ("Recipient Agency", "recipient_agency"),
        ("Recipient Type", "recipient_type"),
        ("Classification", "classification"),
        ("Subject Summary", "subject_summary"),
        ("Linked Reports", "linked_report_count"),
        ("Linked Entities", "linked_entity_count"),
        ("Linked Cases", "linked_case_count"),
        ("Organization", "org_name"),
        ("Disseminated At", "disseminated_at"),
    ]
    _write_sheet(ws, "Disseminations", columns, records)
    return _finalize(wb)


def build_entities_xlsx(records: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    columns: list[tuple[str, str]] = [
        ("Entity ID", "id"),
        ("Type", "entity_type"),
        ("Display Value", "display_value"),
        ("Display Name", "display_name"),
        ("Risk Score", "risk_score"),
        ("Severity", "severity"),
        ("Reports", "report_count"),
        ("Reporting Orgs", "reporting_orgs_count"),
        ("Total Exposure", "total_exposure"),
        ("Status", "status"),
    ]
    _write_sheet(ws, "Entities", columns, records)
    return _finalize(wb)
