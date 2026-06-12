"""Parse an uploaded .xlsx workbook into the same row shape as the CSV parser.

Returns ``list[dict[str, str]]`` mirroring ``csv.DictReader`` output: the first
non-empty row is the header, each subsequent row becomes a dict keyed by header
name with stringified cell values. Date/datetime cells are emitted as ISO 8601
strings so the downstream ``_parse_timestamp`` (datetime.fromisoformat) accepts
them. Reads only the first worksheet.

openpyxl is already a dependency (used by ``services/xlsx_export``). We open in
read-only + data_only mode so formula cells yield their cached values and large
files stream rather than load wholesale.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        # Excel TRUE/FALSE — keep as lowercase string, not "True"/"False".
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        # 1000.0 -> "1000" so account numbers / ints don't pick up a .0 tail.
        return str(int(value))
    return str(value).strip()


def parse_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            return []

        rows_iter = worksheet.iter_rows(values_only=True)

        # First row with any non-empty cell is the header.
        header: list[str] | None = None
        for raw in rows_iter:
            cells = [_cell_to_str(v) for v in raw]
            if any(cells):
                header = cells
                break
        if header is None:
            return []

        results: list[dict[str, str]] = []
        for raw in rows_iter:
            cells = [_cell_to_str(v) for v in raw]
            if not any(cells):
                continue  # skip fully-blank rows (trailing rows are common)
            row: dict[str, str] = {}
            for idx, key in enumerate(header):
                if not key:
                    continue  # ignore unnamed columns
                row[key] = cells[idx] if idx < len(cells) else ""
            results.append(row)
        return results
    finally:
        workbook.close()
